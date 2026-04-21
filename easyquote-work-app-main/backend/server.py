from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
from datetime import datetime, timezone, date
from io import BytesIO

# ReportLab imports
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI(title="EasyQuote Work", version="1.0.0")

# Create router with /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============= MODELS =============

class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(datetime.now(timezone.utc).timestamp()).replace('.', ''))
    name: str = Field(..., min_length=1, max_length=255)
    phone: str = Field(..., min_length=1, max_length=50)
    address: str = Field(..., min_length=1, max_length=500)
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    phone: str = Field(..., min_length=1, max_length=50)
    address: str = Field(..., min_length=1, max_length=500)
    notes: Optional[str] = None

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None

class JobType(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(datetime.now(timezone.utc).timestamp()).replace('.', ''))
    name: str = Field(..., min_length=1, max_length=255)
    unit: str = Field(..., min_length=1, max_length=50)
    price_per_unit: float = Field(..., ge=0)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class JobTypeCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    unit: str = Field(..., min_length=1, max_length=50)
    price_per_unit: float = Field(..., ge=0)

class JobTypeUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    price_per_unit: Optional[float] = None

class QuoteLineItem(BaseModel):
    job_type_id: str
    job_name: str
    unit: str
    quantity: float = Field(..., gt=0)
    price_per_unit: float = Field(..., ge=0)
    total: float = Field(..., ge=0)

class Quote(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(datetime.now(timezone.utc).timestamp()).replace('.', ''))
    quote_number: str
    customer_id: str
    customer_name: str
    customer_phone: str
    customer_address: str
    line_items: List[QuoteLineItem]
    subtotal: float
    iva: float
    total: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class QuoteCreate(BaseModel):
    customer_id: str
    line_items: List[QuoteLineItem] = Field(..., min_length=1)

class WorkReport(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(datetime.now(timezone.utc).timestamp()).replace('.', ''))
    work_date: date
    customer_id: str
    customer_name: str
    job_site: str
    job_description: str
    hours_worked: float = Field(..., gt=0)
    earned_amount: float = Field(..., ge=0)
    hourly_rate: float
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class WorkReportCreate(BaseModel):
    work_date: date
    customer_id: str
    job_site: str
    job_description: str
    hours_worked: float = Field(..., gt=0)
    earned_amount: float = Field(..., ge=0)
    notes: Optional[str] = None

class WorkReportUpdate(BaseModel):
    work_date: Optional[date] = None
    customer_id: Optional[str] = None
    job_site: Optional[str] = None
    job_description: Optional[str] = None
    hours_worked: Optional[float] = None
    earned_amount: Optional[float] = None
    notes: Optional[str] = None

class CompanySettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default="company_settings")
    company_name: str = Field(..., min_length=1, max_length=255)
    owner_name: str = Field(..., min_length=1, max_length=255)
    vat_number: str = Field(..., min_length=1, max_length=50)
    tax_code: str = Field(..., min_length=1, max_length=50)
    address: str = Field(..., min_length=1, max_length=500)
    phone: str = Field(..., min_length=1, max_length=50)
    email: str = Field(..., min_length=1, max_length=100)
    logo_base64: Optional[str] = None
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CompanySettingsUpdate(BaseModel):
    company_name: Optional[str] = None
    owner_name: Optional[str] = None
    vat_number: Optional[str] = None
    tax_code: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    logo_base64: Optional[str] = None

# ============= PDF GENERATION =============

def generate_quote_pdf(quote: dict, company: dict) -> bytes:
    """Genera PDF preventivo con ReportLab."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    verde = colors.HexColor('#1B3A24')
    arancio = colors.HexColor('#D56F53')

    style_title = ParagraphStyle('title', fontSize=20, textColor=verde, spaceAfter=4, fontName='Helvetica-Bold')
    style_sub = ParagraphStyle('sub', fontSize=9, textColor=colors.grey, spaceAfter=2)
    style_h2 = ParagraphStyle('h2', fontSize=11, textColor=verde, spaceBefore=12, spaceAfter=4, fontName='Helvetica-Bold')
    style_normal = ParagraphStyle('normal', fontSize=9, spaceAfter=2)
    style_small = ParagraphStyle('small', fontSize=8, textColor=colors.grey)

    elements = []

    # Header azienda
    elements.append(Paragraph(company.get('company_name', ''), style_title))
    elements.append(Paragraph(company.get('address', ''), style_sub))
    elements.append(Paragraph(f"Tel: {company.get('phone', '')} | Email: {company.get('email', '')}", style_sub))
    elements.append(Paragraph(f"P.IVA: {company.get('vat_number', '')} | C.F.: {company.get('tax_code', '')}", style_sub))
    elements.append(Spacer(1, 0.5*cm))

    # Linea separatrice
    elements.append(Table([['']], colWidths=[17*cm],
        style=TableStyle([('LINEABOVE', (0,0), (-1,-1), 1.5, verde)])))
    elements.append(Spacer(1, 0.3*cm))

    # Numero e data preventivo
    created_at = quote['created_at']
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at)
    quote_date = created_at.strftime("%d/%m/%Y")

    info_data = [
        [
            Paragraph("<b>PREVENTIVO</b>", ParagraphStyle('pv', fontSize=16, textColor=arancio, fontName='Helvetica-Bold')),
            Paragraph(f"<b>N°:</b> {quote['quote_number']}<br/><b>Data:</b> {quote_date}", ParagraphStyle('info', fontSize=10, alignment=TA_RIGHT))
        ]
    ]
    elements.append(Table(info_data, colWidths=[10*cm, 7*cm],
        style=TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')])))
    elements.append(Spacer(1, 0.5*cm))

    # Dati cliente
    elements.append(Paragraph("CLIENTE", style_h2))
    elements.append(Paragraph(f"<b>{quote['customer_name']}</b>", style_normal))
    elements.append(Paragraph(quote['customer_address'], style_normal))
    elements.append(Paragraph(f"Tel: {quote['customer_phone']}", style_normal))
    elements.append(Spacer(1, 0.5*cm))

    # Tabella voci
    elements.append(Paragraph("DETTAGLIO LAVORI", style_h2))

    table_data = [['Descrizione', 'Unità', 'Qtà', '€/Unità', 'Totale']]
    for item in quote['line_items']:
        if isinstance(item, dict):
            job_name = item['job_name']
            unit = item['unit']
            quantity = item['quantity']
            price_per_unit = item['price_per_unit']
            total = item['total']
        else:
            job_name = item.job_name
            unit = item.unit
            quantity = item.quantity
            price_per_unit = item.price_per_unit
            total = item.total

        table_data.append([
            job_name,
            unit,
            f"{quantity:.2f}",
            f"€ {price_per_unit:.2f}",
            f"€ {total:.2f}"
        ])

    table = Table(table_data, colWidths=[7*cm, 2.5*cm, 2*cm, 2.5*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), verde),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F5F5')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.4*cm))

    # Totali
    totali_data = [
        ['', 'Imponibile:', f"€ {quote['subtotal']:.2f}"],
        ['', 'IVA 22%:', f"€ {quote['iva']:.2f}"],
        ['', 'TOTALE:', f"€ {quote['total']:.2f}"],
    ]
    totali_table = Table(totali_data, colWidths=[9*cm, 4*cm, 4*cm])
    totali_table.setStyle(TableStyle([
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (1,2), (-1,2), 'Helvetica-Bold'),
        ('FONTSIZE', (1,2), (-1,2), 11),
        ('TEXTCOLOR', (1,2), (-1,2), verde),
        ('LINEABOVE', (1,2), (-1,2), 1, verde),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(totali_table)

    # Footer
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph("Preventivo valido per 30 giorni dalla data di emissione.", style_small))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()

# ============= CUSTOMER ROUTES =============

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer: CustomerCreate):
    """Crea un nuovo cliente."""
    customer_dict = customer.model_dump()
    customer_obj = Customer(**customer_dict)

    doc = customer_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()

    await db.customers.insert_one(doc)
    return customer_obj

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(search: Optional[str] = None):
    """Ottieni tutti i clienti con ricerca opzionale."""
    query = {}
    if search:
        query = {
            "$or": [
                {"name": {"$regex": search, "$options": "i"}},
                {"phone": {"$regex": search, "$options": "i"}}
            ]
        }

    customers = await db.customers.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)

    for customer in customers:
        if isinstance(customer['created_at'], str):
            customer['created_at'] = datetime.fromisoformat(customer['created_at'])

    return customers

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str):
    """Ottieni un cliente specifico."""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})

    if not customer:
        raise HTTPException(status_code=404, detail="Cliente non trovato")

    if isinstance(customer['created_at'], str):
        customer['created_at'] = datetime.fromisoformat(customer['created_at'])

    return customer

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer_update: CustomerUpdate):
    """Aggiorna un cliente esistente."""
    update_dict = {k: v for k, v in customer_update.model_dump().items() if v is not None}

    if not update_dict:
        raise HTTPException(status_code=400, detail="Nessun campo da aggiornare")

    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": update_dict}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cliente non trovato")

    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if isinstance(customer['created_at'], str):
        customer['created_at'] = datetime.fromisoformat(customer['created_at'])

    return customer

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str):
    """Elimina un cliente."""
    result = await db.customers.delete_one({"id": customer_id})

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cliente non trovato")

    return {"message": "Cliente eliminato con successo"}

# ============= JOB TYPE ROUTES =============

@api_router.post("/job-types", response_model=JobType)
async def create_job_type(job_type: JobTypeCreate):
    """Crea un nuovo tipo di lavoro."""
    job_type_dict = job_type.model_dump()
    job_type_obj = JobType(**job_type_dict)

    doc = job_type_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()

    await db.job_types.insert_one(doc)
    return job_type_obj

@api_router.get("/job-types", response_model=List[JobType])
async def get_job_types():
    """Ottieni tutti i tipi di lavoro."""
    job_types = await db.job_types.find({}, {"_id": 0}).sort("name", 1).to_list(1000)

    for job_type in job_types:
        if isinstance(job_type['created_at'], str):
            job_type['created_at'] = datetime.fromisoformat(job_type['created_at'])

    return job_types

@api_router.get("/job-types/{job_type_id}", response_model=JobType)
async def get_job_type(job_type_id: str):
    """Ottieni un tipo di lavoro specifico."""
    job_type = await db.job_types.find_one({"id": job_type_id}, {"_id": 0})

    if not job_type:
        raise HTTPException(status_code=404, detail="Tipo di lavoro non trovato")

    if isinstance(job_type['created_at'], str):
        job_type['created_at'] = datetime.fromisoformat(job_type['created_at'])

    return job_type

@api_router.put("/job-types/{job_type_id}", response_model=JobType)
async def update_job_type(job_type_id: str, job_type_update: JobTypeUpdate):
    """Aggiorna un tipo di lavoro esistente."""
    update_dict = {k: v for k, v in job_type_update.model_dump().items() if v is not None}

    if not update_dict:
        raise HTTPException(status_code=400, detail="Nessun campo da aggiornare")

    result = await db.job_types.update_one(
        {"id": job_type_id},
        {"$set": update_dict}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tipo di lavoro non trovato")

    job_type = await db.job_types.find_one({"id": job_type_id}, {"_id": 0})
    if isinstance(job_type['created_at'], str):
        job_type['created_at'] = datetime.fromisoformat(job_type['created_at'])

    return job_type

@api_router.delete("/job-types/{job_type_id}")
async def delete_job_type(job_type_id: str):
    """Elimina un tipo di lavoro."""
    result = await db.job_types.delete_one({"id": job_type_id})

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tipo di lavoro non trovato")

    return {"message": "Tipo di lavoro eliminato con successo"}

# ============= QUOTE ROUTES =============

@api_router.post("/quotes", response_model=Quote)
async def create_quote(quote_create: QuoteCreate):
    """Crea un nuovo preventivo."""
    customer = await db.customers.find_one({"id": quote_create.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Cliente non trovato")

    subtotal = sum(item.total for item in quote_create.line_items)
    iva = subtotal * 0.22
    total = subtotal + iva

    count = await db.quotes.count_documents({}) + 1
    quote_number = f"PRV-{datetime.now(timezone.utc).year}-{count:04d}"

    quote_obj = Quote(
        quote_number=quote_number,
        customer_id=quote_create.customer_id,
        customer_name=customer['name'],
        customer_phone=customer['phone'],
        customer_address=customer['address'],
        line_items=quote_create.line_items,
        subtotal=subtotal,
        iva=iva,
        total=total
    )

    doc = quote_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()

    await db.quotes.insert_one(doc)
    return quote_obj

@api_router.get("/quotes", response_model=List[Quote])
async def get_quotes(customer_id: Optional[str] = None, search: Optional[str] = None):
    """Ottieni tutti i preventivi con filtri opzionali."""
    query = {}
    if customer_id:
        query["customer_id"] = customer_id
    if search:
        query["$or"] = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"quote_number": {"$regex": search, "$options": "i"}}
        ]

    quotes = await db.quotes.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)

    for quote in quotes:
        if isinstance(quote['created_at'], str):
            quote['created_at'] = datetime.fromisoformat(quote['created_at'])

    return quotes

@api_router.get("/quotes/{quote_id}", response_model=Quote)
async def get_quote(quote_id: str):
    """Ottieni un preventivo specifico."""
    quote = await db.quotes.find_one({"id": quote_id}, {"_id": 0})

    if not quote:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")

    if isinstance(quote['created_at'], str):
        quote['created_at'] = datetime.fromisoformat(quote['created_at'])

    return quote

@api_router.get("/quotes/{quote_id}/pdf")
async def download_quote_pdf(quote_id: str):
    """Genera e scarica il PDF del preventivo."""
    quote = await db.quotes.find_one({"id": quote_id}, {"_id": 0})

    if not quote:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")

    if isinstance(quote['created_at'], str):
        quote['created_at'] = datetime.fromisoformat(quote['created_at'])

    company_settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})

    if not company_settings:
        company_settings = {
            "company_name": "EasyQuote Work",
            "owner_name": "Titolare",
            "vat_number": "IT00000000000",
            "tax_code": "XXXXXXXXXXX",
            "address": "Via Esempio 1, 00000 Città",
            "phone": "+39 000 0000000",
            "email": "info@azienda.it",
            "logo_base64": None
        }

    pdf_bytes = generate_quote_pdf(quote, company_settings)

    filename = f"Preventivo_{quote['quote_number']}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: str):
    """Elimina un preventivo."""
    result = await db.quotes.delete_one({"id": quote_id})

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Preventivo non trovato")

    return {"message": "Preventivo eliminato con successo"}

# ============= WORK REPORT ROUTES =============

@api_router.post("/work-reports", response_model=WorkReport)
async def create_work_report(report_create: WorkReportCreate):
    """Crea un nuovo report di lavoro."""
    customer = await db.customers.find_one({"id": report_create.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Cliente non trovato")

    hourly_rate = report_create.earned_amount / report_create.hours_worked

    report_obj = WorkReport(
        work_date=report_create.work_date,
        customer_id=report_create.customer_id,
        customer_name=customer['name'],
        job_site=report_create.job_site,
        job_description=report_create.job_description,
        hours_worked=report_create.hours_worked,
        earned_amount=report_create.earned_amount,
        hourly_rate=hourly_rate,
        notes=report_create.notes
    )

    doc = report_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['work_date'] = doc['work_date'].isoformat()

    await db.work_reports.insert_one(doc)
    return report_obj

@api_router.get("/work-reports", response_model=List[WorkReport])
async def get_work_reports(month: Optional[int] = None, year: Optional[int] = None):
    """Ottieni tutti i report di lavoro con filtri opzionali."""
    query = {}

    if month and year:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)

        query["work_date"] = {
            "$gte": start_date.isoformat(),
            "$lt": end_date.isoformat()
        }

    reports = await db.work_reports.find(query, {"_id": 0}).sort("work_date", -1).to_list(1000)

    for report in reports:
        if isinstance(report['created_at'], str):
            report['created_at'] = datetime.fromisoformat(report['created_at'])
        if isinstance(report['work_date'], str):
            report['work_date'] = date.fromisoformat(report['work_date'])

    return reports

@api_router.get("/work-reports/{report_id}", response_model=WorkReport)
async def get_work_report(report_id: str):
    """Ottieni un report specifico."""
    report = await db.work_reports.find_one({"id": report_id}, {"_id": 0})

    if not report:
        raise HTTPException(status_code=404, detail="Report non trovato")

    if isinstance(report['created_at'], str):
        report['created_at'] = datetime.fromisoformat(report['created_at'])
    if isinstance(report['work_date'], str):
        report['work_date'] = date.fromisoformat(report['work_date'])

    return report

@api_router.put("/work-reports/{report_id}", response_model=WorkReport)
async def update_work_report(report_id: str, report_update: WorkReportUpdate):
    """Aggiorna un report esistente."""
    update_dict = {k: v for k, v in report_update.model_dump().items() if v is not None}

    if not update_dict:
        raise HTTPException(status_code=400, detail="Nessun campo da aggiornare")

    if 'hours_worked' in update_dict or 'earned_amount' in update_dict:
        report = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Report non trovato")

        hours = update_dict.get('hours_worked', report['hours_worked'])
        amount = update_dict.get('earned_amount', report['earned_amount'])
        update_dict['hourly_rate'] = amount / hours

    if 'customer_id' in update_dict:
        customer = await db.customers.find_one({"id": update_dict['customer_id']}, {"_id": 0})
        if not customer:
            raise HTTPException(status_code=404, detail="Cliente non trovato")
        update_dict['customer_name'] = customer['name']

    if 'work_date' in update_dict:
        update_dict['work_date'] = update_dict['work_date'].isoformat()

    result = await db.work_reports.update_one(
        {"id": report_id},
        {"$set": update_dict}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Report non trovato")

    report = await db.work_reports.find_one({"id": report_id}, {"_id": 0})
    if isinstance(report['created_at'], str):
        report['created_at'] = datetime.fromisoformat(report['created_at'])
    if isinstance(report['work_date'], str):
        report['work_date'] = date.fromisoformat(report['work_date'])

    return report

@api_router.delete("/work-reports/{report_id}")
async def delete_work_report(report_id: str):
    """Elimina un report."""
    result = await db.work_reports.delete_one({"id": report_id})

    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Report non trovato")

    return {"message": "Report eliminato con successo"}

@api_router.get("/work-reports/summary/monthly")
async def get_monthly_summary(month: Optional[int] = None, year: Optional[int] = None):
    """Ottieni riepilogo mensile ore e guadagni."""
    if not month or not year:
        today = date.today()
        month = today.month
        year = today.year

    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    query = {
        "work_date": {
            "$gte": start_date.isoformat(),
            "$lt": end_date.isoformat()
        }
    }

    reports = await db.work_reports.find(query, {"_id": 0}).to_list(1000)

    total_hours = sum(report['hours_worked'] for report in reports)
    total_earnings = sum(report['earned_amount'] for report in reports)
    average_hourly = total_earnings / total_hours if total_hours > 0 else 0

    return {
        "month": month,
        "year": year,
        "total_hours": round(total_hours, 2),
        "total_earnings": round(total_earnings, 2),
        "average_hourly_rate": round(average_hourly, 2),
        "reports_count": len(reports)
    }

@api_router.get("/work-reports/export/excel")
async def export_work_reports_excel(month: Optional[int] = None, year: Optional[int] = None):
    """Esporta i report di lavoro in Excel."""
    if not month or not year:
        today = date.today()
        month = today.month
        year = today.year

    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    query = {
        "work_date": {
            "$gte": start_date.isoformat(),
            "$lt": end_date.isoformat()
        }
    }

    reports = await db.work_reports.find(query, {"_id": 0}).sort("work_date", -1).to_list(1000)

    for report in reports:
        if isinstance(report['work_date'], str):
            report['work_date'] = date.fromisoformat(report['work_date'])

    total_hours = sum(report['hours_worked'] for report in reports)
    total_earnings = sum(report['earned_amount'] for report in reports)
    average_hourly = total_earnings / total_hours if total_hours > 0 else 0

    wb = Workbook()
    ws = wb.active
    ws.title = f"Report {month:02d}-{year}"

    header_fill = PatternFill(start_color="1B3A24", end_color="1B3A24", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    summary_fill = PatternFill(start_color="D56F53", end_color="D56F53", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:H1')
    ws['A1'] = f"REPORT ORE - {month:02d}/{year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:B3')
    ws['A3'] = "RIEPILOGO MENSILE"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].fill = summary_fill
    ws['A3'].alignment = Alignment(horizontal='center')

    ws['A4'] = "Ore Totali:"
    ws['B4'] = total_hours
    ws['A4'].font = Font(bold=True)
    ws['B4'].number_format = '0.00'

    ws['A5'] = "Guadagno Totale:"
    ws['B5'] = total_earnings
    ws['A5'].font = Font(bold=True)
    ws['B5'].number_format = '€#,##0.00'

    ws['A6'] = "Tariffa Media Oraria:"
    ws['B6'] = average_hourly
    ws['A6'].font = Font(bold=True)
    ws['B6'].number_format = '€#,##0.00'

    headers = ["Data", "Cliente", "Indirizzo", "Descrizione", "Ore", "Importo", "€/Ora", "Note"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    row = 9
    for report in reports:
        ws.cell(row=row, column=1, value=report['work_date'].strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=report['customer_name'])
        ws.cell(row=row, column=3, value=report['job_site'])
        ws.cell(row=row, column=4, value=report['job_description'])
        ws.cell(row=row, column=5, value=report['hours_worked'])
        ws.cell(row=row, column=6, value=report['earned_amount'])
        ws.cell(row=row, column=7, value=report['hourly_rate'])
        ws.cell(row=row, column=8, value=report.get('notes', ''))

        ws.cell(row=row, column=5).number_format = '0.00'
        ws.cell(row=row, column=6).number_format = '€#,##0.00'
        ws.cell(row=row, column=7).number_format = '€#,##0.00'

        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border

        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 30

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"Report_Ore_{month:02d}_{year}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

# ============= COMPANY SETTINGS ROUTES =============

@api_router.get("/company-settings", response_model=CompanySettings)
async def get_company_settings():
    """Ottieni le impostazioni aziendali."""
    settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})

    if not settings:
        return CompanySettings(
            company_name="Nome Azienda",
            owner_name="Nome Titolare",
            vat_number="IT00000000000",
            tax_code="XXXXXXXXXXX",
            address="Via Esempio 1, 00000 Città",
            phone="+39 000 0000000",
            email="info@azienda.it"
        )

    if isinstance(settings['updated_at'], str):
        settings['updated_at'] = datetime.fromisoformat(settings['updated_at'])

    return settings

@api_router.put("/company-settings", response_model=CompanySettings)
async def update_company_settings(settings_update: CompanySettingsUpdate):
    """Aggiorna le impostazioni aziendali."""
    existing = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})

    if existing:
        update_dict = {k: v for k, v in settings_update.model_dump().items() if v is not None}
        update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()

        await db.company_settings.update_one(
            {"id": "company_settings"},
            {"$set": update_dict}
        )

        settings = await db.company_settings.find_one({"id": "company_settings"}, {"_id": 0})
    else:
        settings_obj = CompanySettings(**settings_update.model_dump(exclude_unset=True))
        doc = settings_obj.model_dump()
        doc['updated_at'] = doc['updated_at'].isoformat()

        await db.company_settings.insert_one(doc)
        settings = doc

    if isinstance(settings['updated_at'], str):
        settings['updated_at'] = datetime.fromisoformat(settings['updated_at'])

    return settings

# ============= MAIN =============

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
